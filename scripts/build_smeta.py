#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Смета для заказчика ЖК «Алые Паруса».
3 листа: «Ставки» (единый источник ставки) · «Детали» (построчно, только
формулы) · «Пакеты» (витрина для клиента, тянет итоги формулами).
Меняешь ставку в одной ячейке — пересчитывается ВСЁ.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/Users/mak/cherkesov/Смета_Алые_Паруса.xlsx"

INK = "201B15"; SAND = "E9E0CF"; SCARLET = "B23B2F"; TEAL = "1C5F57"
WHITE = "FBF5EA"; LINE = "CDC2AD"
RUB = '#,##0" ₽"'; HRS = '0" ч"'

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(sz=11, bold=False, color=INK):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def fill(hexc):
    return PatternFill("solid", fgColor=hexc)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
RATE = "'Ставки'!$B$3"

wb = openpyxl.Workbook()

# ══ Ставки ══
rs = wb.active; rs.title = "Ставки"; rs.sheet_view.showGridLines = False
rs["A1"] = "Смета · ЖК «Алые Паруса»"; rs["A1"].font = font(16, True, SCARLET)
rs["A2"] = "Единый источник ставки — меняешь ячейку, пересчитывается вся смета."
rs["A2"].font = font(10, False, "6B6455")
rs["A4"] = "Ставка, ₽/час"; rs["A4"].font = font(12, True)
rs["B4"] = 4500
rs["B4"].font = font(12, True, SCARLET); rs["B4"].number_format = RUB
rs["B4"].fill = fill(SAND); rs["B4"].border = border
# ставка живёт в B3 по ТЗ — держим ссылку на B3; кладём значение и туда
rs["A3"] = "Ставка (используется в формулах, = ниже)"; rs["A3"].font = font(9, False, "6B6455")
rs["B3"] = "=B4"; rs["B3"].font = font(11, True, SCARLET); rs["B3"].number_format = RUB
rs["A6"] = "↑ Меняешь ставку — пересчитываются все пакеты и итоги."
rs["A6"].font = font(10, False, "6B6455")
rs.column_dimensions["A"].width = 34; rs.column_dimensions["B"].width = 16

# ══ Детали ══
d = wb.create_sheet("Детали"); d.sheet_view.showGridLines = False
for col, w in zip("ABCD", (16, 74, 9, 17)):
    d.column_dimensions[col].width = w
for i, h in enumerate(["Тип", "Работа", "Часы", "Стоимость"], 1):
    c = d.cell(1, i, h); c.font = font(11, True, WHITE); c.fill = fill(INK)
    c.alignment = CENTER if i != 2 else LEFT; c.border = border
d.freeze_panes = "A2"

row = 2; subtotal_rows = {}

def section(key, title, items, subtotal_label="Подытог по блоку"):
    global row
    tc = d.cell(row, 1, title); d.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    tc.font = font(12, True, SCARLET); tc.fill = fill(SAND); tc.alignment = LEFT
    for col in range(1, 5): d.cell(row, col).border = border
    row += 1; start = row
    for (t, work, hours) in items:
        d.cell(row, 1, t).font = font(10, color="6B6455")
        d.cell(row, 2, work).font = font(11)
        d.cell(row, 3, hours).number_format = HRS
        d.cell(row, 4, f"=C{row}*{RATE}").number_format = RUB
        d.cell(row, 1).alignment = CENTER; d.cell(row, 2).alignment = LEFT
        d.cell(row, 3).alignment = CENTER; d.cell(row, 4).alignment = RIGHT
        for col in range(1, 5): d.cell(row, col).border = border
        row += 1
    end = row - 1
    d.cell(row, 2, subtotal_label).font = font(11, True)
    d.cell(row, 3, f"=SUM(C{start}:C{end})").number_format = HRS
    d.cell(row, 4, f"=SUM(D{start}:D{end})").number_format = RUB
    for col in range(1, 5):
        cc = d.cell(row, col); cc.fill = fill("FBF9F3"); cc.font = font(11, True); cc.border = border
    d.cell(row, 3).alignment = CENTER; d.cell(row, 4).alignment = RIGHT
    subtotal_rows[key] = row; row += 2

def totals_row(label, formula_d, formula_c=None, strong=False):
    global row
    bg = SCARLET if strong else "FBF9F3"; fg = WHITE if strong else INK
    d.cell(row, 2, label).font = font(12 if strong else 11, True, fg)
    if formula_c:
        d.cell(row, 3, formula_c).number_format = HRS; d.cell(row, 3).alignment = CENTER
    d.cell(row, 4, formula_d).number_format = RUB; d.cell(row, 4).alignment = RIGHT
    for col in range(1, 5):
        cc = d.cell(row, col); cc.fill = fill(bg); cc.font = font(12 if strong else 11, True, fg); cc.border = border
    r = row; row += 2; return r

section("p1", "① Фундамент и инфраструктура проекта", [
    ("Функционал", "Чистая (гексагональная) архитектура: слои domain / application / infrastructure / presentation", 2),
    ("Функционал", "Каркас Next.js 15 (App Router, RSC), TypeScript strict, конфигурация", 1),
    ("Функционал", "Docker: multi-stage сборка, standalone-образ, docker-compose, healthcheck", 2),
    ("Функционал", "Медиа-пайплайн (ffmpeg): нарезка кадров, постеры, фолбэки, оптимизация рендеров", 1),
    ("Функционал", "Юнит-тесты (Vitest) + типобезопасность контрактов", 1),
])
section("p2", "② Дизайн-система", [
    ("Функционал", "Светлая тема «известняковая монография»: палитра, токены, контраст AA", 2),
    ("Функционал", "Типографика: Unbounded + Manrope (кириллица), подключение шрифтов", 1),
    ("Функционал", "Фирменные приёмы: фолио, hairline-линии, «ватерлиния», блок цифр, matted-плиты, curved-quay, зерно, dawn-wipe", 2),
    ("Функционал", "Адаптив + доступность (reduced-motion, фокус, контраст)", 1),
])
section("p3", "③ Hero со скролл-скрабом", [
    ("Функционал", "Движок скролл-скраба: canvas image-sequence, GSAP ScrollTrigger + Lenis, прогрессивная загрузка кадров", 3),
    ("Функционал", "Главы-оверлеи (Гриновская арка), синхрон с кадрами, reduced-motion фолбэк", 1),
    ("Функционал", "Мультивидео + переключатель видео в хедере (общий контекст состояния)", 2),
])
section("p4", "④ Секции сайта", [
    ("Функционал", "Хедер (навигация, поведение при скролле, тумблер) + футер", 1),
    ("Функционал", "«О комплексе» + монументальный блок цифр (табличные цифры)", 1),
    ("Функционал", "Полосы «Надёжность» и «Всё рядом» (эскроу/кадастр/бренды, минуты пешком)", 1),
    ("Функционал", "«Локация» + кинопролёт", 1),
    ("Функционал", "«Инфраструктура» + пролёты + удобства", 1),
    ("Функционал", "«Апартаменты»: отделка + галерея рендеров", 1),
    ("Функционал", "«Инвестиции»: метрики доходности", 1),
    ("Функционал", "Форма заявки + CTA-блок", 1),
])
section("p5", "⑤ Интерактив и анимации", [
    ("Функционал", "Индикатор прогресса, счётчики цифр (count-up), проявление заголовков", 1),
    ("Функционал", "Ховеры карточек, зум галереи, подчёркивания навигации", 1),
    ("Функционал", "Бренд-парус (SVG) в финале героя и у формы брони", 1),
    ("Функционал", "Кинематографичный блок ScrollStory: фон-скраб + выезжающий сбоку текст", 1),
])
section("p6", "⑥ Контроль качества", [
    ("Функционал", "Многоагентные код-ревью (архитектура, гидрация, скраб, доступность) + правки", 2),
])
section("ai", "⑦ Контент, созданный нейросетями", [
    ("AI-контент", "Морф стройки день→ночь (Seedance FLF): ключевые кадры K0–K6 + 6 клипов + сборка", 4),
    ("AI-контент", "Альтернативное видео стройки (диагональный ракурс)", 2),
    ("AI-контент", "3 кинематографичных пролёта (Seedance FLF)", 2),
    ("AI-контент", "Фотореалистичные дневные рендеры: перевод из ночных + удаление брендинга (6 шт)", 2),
    ("AI-контент", "Копирайт/сценарий: Гриновская арка героя, тексты секций", 1),
    ("AI-контент", "Арт-дирекшн: генерация и отбор дизайн-направления, палитры", 1),
])
section("fix", "⑧ Исправления по ходу — включено, 0 ₽ (жест доброй воли)", [
    ("Правка", "Пустой блок в галерее «Апартаменты» — устранён", 0),
    ("Правка", "Сброс героя в стартовое состояние при переключении видео на середине", 0),
    ("Правка", "Контраст неактивной кнопки тумблера над видео", 0),
    ("Правка", "Инерция скролла — защита от проскакивания блоков", 0),
], subtotal_label="Итого правок (не тарифицируется)")

func_sum = "+".join(f"D{subtotal_rows[k]}" for k in ["p1", "p2", "p3", "p4", "p5", "p6"])
func_hrs = "+".join(f"C{subtotal_rows[k]}" for k in ["p1", "p2", "p3", "p4", "p5", "p6"])
r_func = totals_row("ИТОГО: реализованный функционал (①–⑥)", f"={func_sum}", f"={func_hrs}")
r_ai = totals_row("ИТОГО: AI-контент (⑦)", f"=D{subtotal_rows['ai']}", f"=C{subtotal_rows['ai']}")
r_realized = totals_row("ИТОГО РЕАЛИЗОВАНО (функционал + AI-контент)",
                        f"=D{r_func}+D{r_ai}", f"=C{r_func}+C{r_ai}", strong=True)

section("opt", "⑨ Возможные доработки (по запросу, ориентировочно)", [
    ("Доработка", "ROI-калькулятор «Прилив капитала»: интерактивный расчёт доходности → PDF + бронь", 24),
    ("Доработка", "«Парус-нить»: единый алый парус hero→бронь, имя на парусе при отправке", 16),
    ("Доработка", "Drone-шторка «рендер → стройка сегодня»: снятие страха недостроя", 14),
    ("Доработка", "Изохрона «8 минут пешком»: интерактивная карта пешей доступности", 16),
    ("Доработка", "Интерактивный генплан 46 корпусов + подбор квартиры (этаж/вид/статус)", 30),
    ("Доработка", "Диммер «день/ночь», управляемый пользователем", 8),
    ("Доработка", "Полноценная блочная навигация (fullpage-снап)", 12),
    ("Доработка", "Мобильное меню (бургер) + тумблер видео на телефоне", 6),
    ("Доработка", "Бэкенд формы заявок + интеграция с CRM (Bitrix/amoCRM)", 16),
    ("Доработка", "Второй ScrollStory-блок под инвестора (пролёт + цифры доходности)", 6),
    ("Доработка", "Перегенерация морфа стройки в 1080p (финальное качество)", 6),
    ("Доработка", "Интеграция CMS (замена статических данных без переписывания слоёв)", 20),
    ("Доработка", "Мультиязычность (RU / EN)", 16),
])
r_buf = totals_row("Буфер 15% на доработки (неизвестное/интеграции)", f"=D{subtotal_rows['opt']}*0.15", None)
r_opt_total = totals_row("ИТОГО доработки с буфером", f"=D{subtotal_rows['opt']}+D{r_buf}", f"=C{subtotal_rows['opt']}", strong=True)

# ══ Пакеты ══
p = wb.create_sheet("Пакеты"); p.sheet_view.showGridLines = False
for col, w in zip("ABCD", (34, 66, 10, 18)):
    p.column_dimensions[col].width = w
p["A1"] = "ЖК «Алые Паруса» · пакеты работ"; p["A1"].font = font(16, True, SCARLET)
p["A2"] = "Стоимость тянется формулами из «Деталей», ставка — на листе «Ставки»."
p["A2"].font = font(10, False, "6B6455")
for i, h in enumerate(["Пакет", "Что получает заказчик", "Часы", "Стоимость"], 1):
    c = p.cell(4, i, h); c.font = font(11, True, WHITE); c.fill = fill(INK)
    c.alignment = CENTER if i not in (1, 2) else LEFT; c.border = border

def pkg_row(r, name, desc, src_row, strong=False, accent=False):
    bg = SCARLET if strong else (SAND if accent else "FFFFFF"); fg = WHITE if strong else INK
    p.cell(r, 1, name).font = font(12 if strong else 11, True, fg)
    p.cell(r, 2, desc).font = font(10, False, fg if strong else "3B342B")
    p.cell(r, 3, f"=Детали!C{src_row}").number_format = HRS
    p.cell(r, 4, f"=Детали!D{src_row}").number_format = RUB
    p.cell(r, 1).alignment = LEFT; p.cell(r, 2).alignment = LEFT
    p.cell(r, 3).alignment = CENTER; p.cell(r, 4).alignment = RIGHT
    for col in range(1, 5):
        p.cell(r, col).border = border
        if bg != "FFFFFF": p.cell(r, col).fill = fill(bg)
    p.row_dimensions[r].height = 46

pkg_row(5, "① Сайт под ключ", "Премиальный светлый сайт со скролл-«стройкой», кинематографичным блоком, интерактивом и адаптивом. Готов к запуску.", r_func)
pkg_row(6, "② AI-контент", "Видео роста дома день→ночь, 3 пролёта, фотореалистичные дневные рендеры без брендинга, тексты и сценарий.", r_ai)
pkg_row(7, "③ ИТОГО реализовано", "Всё готовое: функционал сайта + весь контент нейросетей. Плюс исправления по ходу — бесплатно.", r_realized, strong=True)
pkg_row(8, "④ Возможные доработки", "Калькулятор доходности, «парус-нить», доказательство стройки, карта пешей доступности, генплан+подбор, CRM и др.", r_opt_total, accent=True)

p.cell(10, 1, "Жест доброй воли: 4 исправления по ходу проекта включены — 0 ₽.").font = font(10, False, TEAL)
p.cell(11, 1, "Реализованное — фактические часы по объёму работ. «Возможные доработки» — по полной почасовке, ориентировочно.").font = font(10, False, "6B6455")
p.cell(12, 1, "Суммы пересчитываются при смене ставки на листе «Ставки».").font = font(10, False, "6B6455")
wb.active = wb.sheetnames.index("Пакеты")
wb.save(OUT)
print("saved:", OUT)
